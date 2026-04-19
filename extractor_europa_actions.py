#!/usr/bin/env python3
"""
extractor_europa_actions.py  –  Hipopótamo Europa S.L.
Flujo: Gmail IMAP → Excel ERP → data.json (GitHub Pages)
"""
import os, re, json, imaplib, email, tempfile, traceback
from email.header import decode_header
from email.utils import parsedate_to_datetime
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo
import openpyxl

# ── Deps opcionales (Google Sheets / Drive) ────────────────────────────────────
try:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    HAS_GOOGLE = True
except ImportError:
    HAS_GOOGLE = False

# ── Festivos 2025-2026 (nacionales España + Aragón + Lleida) ──────────────────
FESTIVOS_ZARAGOZA = {
    "2025-01-01","2025-01-06","2025-04-17","2025-04-18",
    "2025-04-23","2025-05-01","2025-10-12","2025-11-01",
    "2025-12-06","2025-12-08","2025-12-25",
    "2026-01-01","2026-01-06","2026-04-02","2026-04-03",
    "2026-04-23","2026-05-01","2026-10-12","2026-11-01",
    "2026-12-07","2026-12-08","2026-12-25",
}
FESTIVOS_LLEIDA = {
    "2025-01-01","2025-01-06","2025-04-17","2025-04-18",
    "2025-05-01","2025-06-24","2025-09-11","2025-10-12",
    "2025-11-01","2025-12-08","2025-12-25","2025-12-26",
    "2026-01-01","2026-01-06","2026-04-02","2026-04-03",
    "2026-05-01","2026-06-24","2026-09-11","2026-10-12",
    "2026-11-01","2026-12-08","2026-12-25","2026-12-26",
}

# ── Canal → Centro mapping ────────────────────────────────────────────────────
CANAL_MAP = {
    "ZARAGOZA":         "CENTRAL",
    "CENTRAL":          "CENTRAL",
    "LERIDA":           "ALCARRAS",
    "LÉRIDA":           "ALCARRAS",
    "ALCARRAS":         "ALCARRAS",
    "ALMOZARA":         "ALMOZARA",
    "CORONA DE ARAGON": "CORONA",
    "CORONA DE ARAGÓN": "CORONA",
    "CORONA":           "CORONA",
}

# ── Presupuesto 2026 por mes y centro ─────────────────────────────────────────
BUDGET_26 = {
    1:  {"CENTRAL":139000,"ALCARRAS":179000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":437000},
    2:  {"CENTRAL":129000,"ALCARRAS":197000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":445000},
    3:  {"CENTRAL":129000,"ALCARRAS":179000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":427000},
    4:  {"CENTRAL":129000,"ALCARRAS":179000,"ALMOZARA":29000,"CORONA":80000,"TOTAL":417000},
    5:  {"CENTRAL":129000,"ALCARRAS":189000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":437000},
    6:  {"CENTRAL":109000,"ALCARRAS":179000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":407000},
    7:  {"CENTRAL":109000,"ALCARRAS":179000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":407000},
    8:  {"CENTRAL":109000,"ALCARRAS":179000,"ALMOZARA":29000,"CORONA":80000,"TOTAL":397000},
    9:  {"CENTRAL":129000,"ALCARRAS":219000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":467000},
    10: {"CENTRAL":129000,"ALCARRAS":219000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":467000},
    11: {"CENTRAL":149000,"ALCARRAS":179000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":447000},
    12: {"CENTRAL":109000,"ALCARRAS":159000,"ALMOZARA":39000,"CORONA":80000,"TOTAL":387000},
}

# ── Datos históricos estáticos (reales confirmados, en euros) ─────────────────
HIST_CENTRAL = {
    "2021":[127000,94000,117000,109000,85000,129000,104000,118000,105000,109000,156000,79000],
    "2022":[132000,102000,110000,109000,99000,96000,100000,117000,103000,131000,156000,79000],
    "2023":[88000,123000,113000,127000,119000,80000,76000,95000,103000,106000,132000,87000],
    "2024":[101506,134080,123046,101291,129094,68234,84369,86137,101560,89058,146000,80000],
    "2025":[134241,90429,100762,101291,91000,59000,87000,85000,111000,108000,109000,80000],
}
HIST_ALCARRAS = {
    "2021":[74000,23000,134000,141000,158000,158000,150000,148000,171000,195000,129000,126000],
    "2022":[135134,123538,129766,168079,145953,144920,109836,174999,151372,151996,128876,126322],
    "2023":[123158,175111,103496,115414,154750,158397,175256,122980,187436,188734,198200,114878],
    "2024":[141770,191828,128271,130955,158339,148447,182127,141899,146927,162587,197000,145000],
    "2025":[158490,182979,149535,167279,174381,143252,169986,151958,199832,197779,141503,111000],
}
HIST_ALMOZARA = {
    "2024":[22239,23363,35191,20000,28000,27000,19000,10000,51000,52000,36000,25000],
    "2025":[34694,32289,28364,19000,24000,26000,20000,14000,24000,21000,22000,15000],
}
HIST_CORONA = {
    "2025":[30504,86471,60831,39462,56857,49123,48907,29502,77502,58752,50845,31172],
}

# ── Datos mensuales 2025 y 2026 confirmados por centro ────────────────────────
MONTHLY_BY_CENTER = {
    "2025": {
        1:{"CENTRAL":134241,"ALCARRAS":158490,"ALMOZARA":34694,"CORONA":30504},
        2:{"CENTRAL":90429, "ALCARRAS":182979,"ALMOZARA":32289,"CORONA":86471},
        3:{"CENTRAL":100762,"ALCARRAS":149535,"ALMOZARA":28364,"CORONA":60831},
        4:{"CENTRAL":101291,"ALCARRAS":167279,"ALMOZARA":19000,"CORONA":39462},
        5:{"CENTRAL":91000, "ALCARRAS":174381,"ALMOZARA":24000,"CORONA":56857},
        6:{"CENTRAL":59000, "ALCARRAS":143252,"ALMOZARA":26000,"CORONA":49123},
        7:{"CENTRAL":87000, "ALCARRAS":169986,"ALMOZARA":20000,"CORONA":48907},
        8:{"CENTRAL":85000, "ALCARRAS":151958,"ALMOZARA":14000,"CORONA":29502},
        9:{"CENTRAL":111000,"ALCARRAS":199832,"ALMOZARA":24000,"CORONA":77502},
        10:{"CENTRAL":108000,"ALCARRAS":197779,"ALMOZARA":21000,"CORONA":58752},
        11:{"CENTRAL":109000,"ALCARRAS":141503,"ALMOZARA":22000,"CORONA":50845},
        12:{"CENTRAL":80000, "ALCARRAS":111000,"ALMOZARA":15000,"CORONA":31172},
    },
    "2026": {
        1:{"CENTRAL":104424,"ALCARRAS":194066,"ALMOZARA":23491,"CORONA":40286},
        2:{"CENTRAL":122057,"ALCARRAS":222208,"ALMOZARA":31791,"CORONA":38567},
        3:{"CENTRAL":107997,"ALCARRAS":141522,"ALMOZARA":20950,"CORONA":34211},
    },
}

TOTAL_MONTHLY = {
    2023:[202945,297824,216672,242818,274082,238758,251016,217911,290767,294400,329986,202116],
    2024:[265515,350421,286508,252700,315800,246211,285907,245262,299321,305765,391524,256428],
    2025:[357929,392168,339492,322682,346288,278616,326300,280499,411993,385315,323646,248123],
}


# ─────────────────────────────────────────────────────────────────────────────
#  CALENDARIO - Días laborables (Lun-Sáb) con festivos
# ─────────────────────────────────────────────────────────────────────────────
def contar_dias_laborables(desde: date, hasta: date, festivos: set) -> int:
    """Cuenta días Lun-Sáb que no sean festivos, desde <= hasta."""
    count = 0
    d = desde
    while d <= hasta:
        if d.weekday() != 6:  # 6 = domingo
            if d.strftime("%Y-%m-%d") not in festivos:
                count += 1
        d += timedelta(days=1)
    return count


def get_commercial_month(ref_date: date):
    """
    Devuelve (year, month) del mes comercial al que pertenece ref_date.
    Mes comercial: del 26 del mes anterior al 25 del mes actual.
    """
    if ref_date.day >= 26:
        # Entramos en el siguiente mes comercial
        if ref_date.month == 12:
            return ref_date.year + 1, 1
        return ref_date.year, ref_date.month + 1
    return ref_date.year, ref_date.month


def get_commercial_period(com_year: int, com_month: int):
    """Devuelve (inicio, fin) del mes comercial (inicio=26 del mes anterior)."""
    if com_month == 1:
        inicio = date(com_year - 1, 12, 26)
    else:
        inicio = date(com_year, com_month - 1, 26)
    fin = date(com_year, com_month, 25)
    return inicio, fin


# ─────────────────────────────────────────────────────────────────────────────
#  IMAP – Búsqueda y descarga del adjunto ERP
# ─────────────────────────────────────────────────────────────────────────────
def imap_connect():
    user = os.environ["GMAIL_USER"]
    pwd  = os.environ["GMAIL_PASSWORD"]
    host = os.environ.get("IMAP_HOST", "imap.gmail.com")
    M = imaplib.IMAP4_SSL(host)
    M.login(user, pwd)
    return M


def decode_str(s):
    parts = decode_header(s)
    result = ""
    for part, enc in parts:
        if isinstance(part, bytes):
            result += part.decode(enc or "utf-8", errors="replace")
        else:
            result += str(part)
    return result


TZ_MADRID = ZoneInfo("Europe/Madrid")
# Meses en inglés para criterio IMAP SINCE (RFC 3501)
_IMAP_MON = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


def _imap_since_str(d: date) -> str:
    return f"{d.day}-{_IMAP_MON[d.month - 1]}-{d.year}"


def _fecha_referencia_correo() -> tuple[date, str]:
    """
    Día civil en Madrid que debe coincidir con la cabecera Date del correo.
    CORREO_REFERENCIA: ayer (defecto), hoy, latest (sin filtrar por día).
    """
    mode = (os.environ.get("CORREO_REFERENCIA", "ayer") or "ayer").strip().lower()
    now = datetime.now(TZ_MADRID)
    if mode in ("latest", "mas_reciente", "reciente"):
        return now.date(), "latest"
    if mode in ("hoy", "today"):
        return now.date(), "hoy"
    return (now - timedelta(days=1)).date(), "ayer"


def _fecha_cabecera_date_madrid(msg: email.message.Message) -> date | None:
    raw = msg.get("Date")
    if not raw:
        return None
    try:
        dt = parsedate_to_datetime(raw)
    except (TypeError, ValueError):
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC"))
    return dt.astimezone(TZ_MADRID).date()


def find_excel_attachment(msg):
    """Extrae el primer adjunto .xlsx del mensaje."""
    for part in msg.walk():
        ct = part.get_content_type()
        cd = part.get("Content-Disposition", "")
        fn = part.get_filename()
        if fn:
            fn = decode_str(fn)
        is_excel = (
            fn and fn.lower().endswith(".xlsx")
        ) or ct in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        )
        if is_excel and fn:
            return part.get_payload(decode=True), fn
    return None, None


def fetch_latest_erp_excel():
    """
    Busca el correo ERP (por defecto: cabecera Date en Madrid = ayer),
    descarga el adjunto y devuelve
    (bytes_excel, filename, uid, fecha_generacion_desde_fichero).

    CORREO_REFERENCIA=ayer|hoy|latest  (por defecto ayer)
    """
    asunto_filtro = os.environ.get("ASUNTO_FILTRO", "Informe_ventas")
    remitente     = os.environ.get("REMITENTE", "")
    ref_dia, ref_mode = _fecha_referencia_correo()
    since_imap = _imap_since_str(ref_dia - timedelta(days=14))

    M = imap_connect()
    M.select("INBOX")

    parts = [f"SINCE {since_imap}"]
    if remitente:
        parts.append(f'FROM "{remitente}"')
    if asunto_filtro:
        parts.append(f'SUBJECT "{asunto_filtro}"')
    search_str = "(" + " ".join(parts) + ")"

    status, data = M.search(None, search_str)
    if status != "OK" or not data[0]:
        status, data = M.search(None, f'(SUBJECT "ventas" SINCE {since_imap})')

    uids = data[0].split() if data[0] else []
    if not uids:
        M.logout()
        raise RuntimeError("No se encontró correo ERP en INBOX (revisar ASUNTO_FILTRO / REMITENTE)")

    def candidato_desde_uid(uid, filtrar_por_dia: date | None):
        status, msg_data = M.fetch(uid, "(RFC822)")
        if status != "OK" or not msg_data or not msg_data[0]:
            return None
        chunk = msg_data[0]
        raw = chunk[1] if isinstance(chunk, tuple) else chunk
        if not isinstance(raw, bytes):
            return None
        msg = email.message_from_bytes(raw)
        if filtrar_por_dia is not None:
            fd = _fecha_cabecera_date_madrid(msg)
            if fd is None or fd != filtrar_por_dia:
                return None
        payload, fname = find_excel_attachment(msg)
        if payload is None:
            return None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
                tf.write(payload)
                tf_path = tf.name
            fecha_gen = extraer_fecha_generacion(tf_path)
            os.unlink(tf_path)
        except Exception:
            fecha_gen = date(2000, 1, 1)
        return (payload, fname, uid, fecha_gen)

    # Más recientes primero (UID mayor al final de la lista típica IMAP)
    tail = uids[-80:] if len(uids) > 80 else uids
    ordered = list(reversed(tail))

    best = None  # (fecha_gen, uid_int, payload, fname, uid)
    filtro = None if ref_mode == "latest" else ref_dia

    if filtro is not None:
        print(f"  Buscando correo con Date (Madrid) = {filtro} (modo {ref_mode})")
        for uid in ordered:
            c = candidato_desde_uid(uid, filtro)
            if c is None:
                continue
            payload, fname, u, fecha_gen = c
            uid_int = int(u)
            cand = (fecha_gen, uid_int, payload, fname, u)
            if best is None or cand[1] > best[1] or (cand[1] == best[1] and cand[0] >= best[0]):
                best = cand

    if best is None and filtro is not None:
        print(f"  ⚠ Ningún correo con Date Madrid = {filtro}; usando el Excel más reciente de la bandeja.")
        for uid in ordered:
            c = candidato_desde_uid(uid, None)
            if c is None:
                continue
            payload, fname, u, fecha_gen = c
            uid_int = int(u)
            cand = (fecha_gen, uid_int, payload, fname, u)
            if best is None or cand[0] > best[0] or (cand[0] == best[0] and cand[1] > best[1]):
                best = cand
    elif best is None:
        for uid in ordered:
            c = candidato_desde_uid(uid, None)
            if c is None:
                continue
            payload, fname, u, fecha_gen = c
            uid_int = int(u)
            cand = (fecha_gen, uid_int, payload, fname, u)
            if best is None or cand[0] > best[0] or (cand[0] == best[0] and cand[1] > best[1]):
                best = cand

    M.logout()
    if best is None:
        raise RuntimeError("No se pudo descargar adjunto Excel del correo ERP")
    fecha_gen, _uid_int, payload, fname, uid = best
    return payload, fname, uid, fecha_gen


# ─────────────────────────────────────────────────────────────────────────────
#  PARSER Excel ERP
# ─────────────────────────────────────────────────────────────────────────────
def extraer_fecha_generacion(path: str) -> date:
    """Lee la fecha de generación del ERP (fila 2: 'Fecha: DD/MM/YY Hora:...')"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        for cell in row:
            if cell and isinstance(cell, str) and "Fecha:" in cell:
                m = re.search(r"Fecha:\s*(\d{1,2})/(\d{1,2})/(\d{2,4})", cell)
                if m:
                    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    if y < 100:
                        y += 2000
                    wb.close()
                    return date(y, mo, d)
    wb.close()
    return date.today()


def extraer_rango_fechas(ws):
    """
    Lee 'Fecha desde: DD/MM/YY   Fecha hasta: DD/MM/YY' del encabezado ERP.
    Devuelve (fecha_desde, fecha_hasta) como objetos date.
    """
    for row in ws.iter_rows(max_row=6, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "Fecha desde:" in cell:
                m = re.findall(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", cell)
                if len(m) >= 2:
                    def parse(t):
                        d, mo, y = int(t[0]), int(t[1]), int(t[2])
                        if y < 100:
                            y += 2000
                        return date(y, mo, d)
                    return parse(m[0]), parse(m[1])
    return None, None


def parsear_excel_erp(path: str):
    """
    Parsea el Excel ERP y devuelve:
      {
        "fecha_generacion": date,
        "fecha_desde": date,
        "fecha_hasta": date,
        "por_canal": {"CENTRAL": float, "ALCARRAS": float, ...},
        "por_vendedor": {"CENTRAL": [{"name":str,"real":float},...], ...},
        "total": float,
      }
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws_name = None
    for sn in wb.sheetnames:
        if sn.upper() in ("VENTAS", "DATOS", "DATOS-AUTOMATIZADA"):
            ws_name = sn
            break
    ws = wb[ws_name] if ws_name else wb.active

    fecha_gen = extraer_fecha_generacion(path)
    fecha_desde, fecha_hasta = extraer_rango_fechas(ws)

    # ── Sección 1: ventas por tienda/canal ────────────────────────────────────
    por_canal_raw = {}   # canal_erp → total
    por_vendedor_raw = {}  # (nombre, canal_erp) → total

    section = None
    for row in ws.iter_rows(values_only=True):
        row = [c for c in row]
        if not any(c for c in row):
            continue
        # Detectar cabeceras de sección
        first = str(row[0] or "").strip()
        if "pedidos de venta por canal" in first.lower():
            section = "canal"
            continue
        if "ventas por vendedor" in first.lower():
            section = "vendedor"
            continue
        if first in ("Tienda", "Vendedor"):
            continue

        if section == "canal":
            # Filas de tienda: ('TZ', 'Tienda Zaragoza', 'CANAL', total)
            # Filas totales: ('Total tienda XX', ...)
            if len(row) >= 4 and row[2] and not str(row[0]).startswith("Total"):
                canal = str(row[2]).strip().upper()
                total = row[3] if isinstance(row[3], (int, float)) else 0
                if total:
                    por_canal_raw[canal] = por_canal_raw.get(canal, 0) + total

        elif section == "vendedor":
            # Filas: ('123', 'Nombre Vendedor', 'CANAL', total)
            if len(row) >= 4 and row[1] and not str(row[0]).startswith("Total"):
                nombre = str(row[1]).strip()
                canal  = str(row[2] or "").strip().upper()
                total  = row[3] if isinstance(row[3], (int, float)) else 0
                if total and canal and not nombre.startswith("TOTAL"):
                    key = (nombre, canal)
                    por_vendedor_raw[key] = por_vendedor_raw.get(key, 0) + total

    wb.close()

    # ── Mapear canal → centro ─────────────────────────────────────────────────
    por_canal = {"CENTRAL":0.0,"ALCARRAS":0.0,"ALMOZARA":0.0,"CORONA":0.0}
    for canal_erp, val in por_canal_raw.items():
        centro = CANAL_MAP.get(canal_erp)
        if centro:
            por_canal[centro] += val

    # ── Agrupar vendedores por centro ─────────────────────────────────────────
    vend_agg = {}  # centro → nombre → total
    for (nombre, canal_erp), val in por_vendedor_raw.items():
        centro = CANAL_MAP.get(canal_erp)
        if not centro:
            continue
        vend_agg.setdefault(centro, {})
        vend_agg[centro][nombre] = vend_agg[centro].get(nombre, 0) + val

    por_vendedor = {}
    for centro, vendedores in vend_agg.items():
        por_vendedor[centro] = [
            {"name": n, "real": round(v, 2)}
            for n, v in sorted(vendedores.items(), key=lambda x: -x[1])
            if v > 0
        ]

    return {
        "fecha_generacion": fecha_gen,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "por_canal": por_canal,
        "por_vendedor": por_vendedor,
        "total": sum(por_canal.values()),
    }


# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE SHEETS – Append histórico (opcional)
# ─────────────────────────────────────────────────────────────────────────────
def get_sheets_service():
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sa_json or not HAS_GOOGLE:
        return None
    import json as _json
    info = _json.loads(sa_json)
    creds = Credentials.from_service_account_info(
        info,
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    return build("sheets", "v4", credentials=creds)


def append_to_sheets(erp_data: dict):
    """Guarda fila en Google Sheets (idempotente por report_key)."""
    sheet_id = os.environ.get("DRIVE_SPREADSHEET_ID", "")
    if not sheet_id:
        return
    svc = get_sheets_service()
    if not svc:
        return
    tab = "Datos"
    fd = erp_data.get("fecha_desde")
    report_key = f"{fd}_{erp_data['fecha_generacion']}" if fd else str(erp_data["fecha_generacion"])
    # Leer claves existentes para idempotencia
    try:
        result = svc.spreadsheets().values().get(
            spreadsheetId=sheet_id, range=f"{tab}!A:A"
        ).execute()
        existing_keys = [r[0] for r in result.get("values", []) if r]
        if report_key in existing_keys:
            print(f"  Sheets: ya existe {report_key}, skip")
            return
    except Exception as e:
        print(f"  Sheets warning: {e}")
    canal = erp_data["por_canal"]
    row = [
        report_key,
        str(erp_data["fecha_generacion"]),
        str(fd or ""),
        str(erp_data.get("fecha_hasta") or ""),
        round(canal.get("CENTRAL", 0), 2),
        round(canal.get("ALCARRAS", 0), 2),
        round(canal.get("ALMOZARA", 0), 2),
        round(canal.get("CORONA", 0), 2),
        round(erp_data["total"], 2),
    ]
    try:
        svc.spreadsheets().values().append(
            spreadsheetId=sheet_id,
            range=f"{tab}!A:I",
            valueInputOption="USER_ENTERED",
            body={"values": [row]},
        ).execute()
        print(f"  Sheets: append OK → {report_key}")
    except Exception as e:
        print(f"  Sheets error: {e}")


# ─────────────────────────────────────────────────────────────────────────────
#  GENERACIÓN data.json
# ─────────────────────────────────────────────────────────────────────────────
def build_rolling(hist_total: dict) -> list:
    """Calcula rolling 12m mensual desde todos los datos históricos disponibles."""
    all_months = []
    for year in sorted(hist_total.keys()):
        months = hist_total[year]
        for mi, val in enumerate(months):
            if val is not None:
                all_months.append((year, mi + 1, val))

    rolling = []
    for i in range(11, len(all_months)):
        window = all_months[i - 11 : i + 1]
        val = sum(x[2] for x in window)
        y, m, _ = all_months[i]
        rolling.append({"label": f"{y}-{m:02d}", "value": round(val)})
    return rolling


def build_hist_total(cur_year: int, cur_month: int, centro_by_year: dict) -> dict:
    """Construye histórico total fusionando estáticos + mes actual."""
    static_totals = {
        2023: [202945,297824,216672,242818,274082,238758,251016,217911,290767,294400,329986,202116],
        2024: [265515,350421,286508,252700,315800,246211,285907,245262,299321,305765,391524,256428],
        2025: [357929,392168,339492,322682,346288,278616,326300,280499,411993,385315,323646,248123],
    }
    # Año 2021 y 2022 estimados
    static_totals[2021] = [201000,117000,251000,250000,243000,287000,254000,266000,276000,304000,285000,205000]
    static_totals[2022] = [267021,225340,239287,276818,245032,240474,209842,292090,254135,283067,285358,205412]

    result = {}
    for year, monthly in sorted(static_totals.items()):
        result[year] = list(monthly)

    # Actualizar/añadir el año actual con datos frescos
    if cur_year not in result:
        result[cur_year] = [None] * 12
    # Rellenar meses completados del año actual desde MONTHLY_BY_CENTER
    for m_idx in range(12):
        m = m_idx + 1
        completed_key = str(cur_year)
        mc = MONTHLY_BY_CENTER.get(completed_key, {}).get(m)
        if mc:
            result[cur_year][m_idx] = round(sum(mc.values()))

    # Mes comercial actual en curso (parcial)
    cur_data = centro_by_year.get(cur_year, {}).get(cur_month)
    if cur_data:
        result[cur_year][cur_month - 1] = round(sum(cur_data.values()))

    return {str(k): v for k, v in sorted(result.items())}


def build_monthly_series(cur_year: int, cur_month: int, real_cur: dict,
                          proj_cur: float) -> dict:
    """Construye series mensuales para el frontend."""
    series = {}

    # Año actual
    months_2026 = []
    completed = MONTHLY_BY_CENTER.get(str(cur_year), {})
    budget_year = BUDGET_26 if cur_year == 2026 else {}

    for m in range(1, 13):
        bc = budget_year.get(m, {})
        if m < cur_month and m in completed:
            rc = completed[m]
            t_real = round(sum(rc.values()))
            months_2026.append({
                "month": m,
                "real": t_real,
                "budget": bc.get("TOTAL", 0),
                "projection": t_real,
                "byCenter": {c: {"real": round(v), "budget": bc.get(c, 0)} for c, v in rc.items()},
            })
        elif m == cur_month:
            months_2026.append({
                "month": m,
                "real": round(real_cur.get("TOTAL", 0), 2),
                "budget": bc.get("TOTAL", 0),
                "projection": round(proj_cur, 2),
                "byCenter": {c: {"real": round(real_cur.get(c, 0), 2), "budget": bc.get(c, 0)}
                             for c in ["CENTRAL","ALCARRAS","ALMOZARA","CORONA"]},
            })
        else:
            months_2026.append({
                "month": m,
                "real": None,
                "budget": bc.get("TOTAL", 0),
                "projection": None,
            })
    series[str(cur_year)] = months_2026

    # Año anterior
    prev_year = cur_year - 1
    prev_monthly = MONTHLY_BY_CENTER.get(str(prev_year), {})
    months_prev = []
    for m in range(1, 13):
        rc = prev_monthly.get(m, {})
        t = round(sum(rc.values())) if rc else TOTAL_MONTHLY.get(prev_year, [None]*12)[m-1]
        months_prev.append({"month": m, "real": t, "budget": 0, "projection": t})
    series[str(prev_year)] = months_prev

    # 2024, 2023
    for yr in [2024, 2023]:
        totals = TOTAL_MONTHLY.get(yr, [])
        series[str(yr)] = [{"month": m+1, "real": v, "budget": 0, "projection": v}
                           for m, v in enumerate(totals)]
    return series


def generar_data_json(erp_data: dict, output_path: str = "data.json"):
    """Construye y escribe data.json completo."""

    fg    = erp_data["fecha_generacion"]
    fd    = erp_data.get("fecha_desde") or date.today()
    fh    = erp_data.get("fecha_hasta") or date.today()
    real  = erp_data["por_canal"]
    vend  = erp_data["por_vendedor"]
    total = erp_data["total"]

    # Determinar mes comercial
    com_year, com_month = get_commercial_month(fd)

    # Días laborables
    inicio_com, fin_com = get_commercial_period(com_year, com_month)
    festivos = FESTIVOS_LLEIDA if False else FESTIVOS_ZARAGOZA  # Zaragoza por defecto
    days_total   = contar_dias_laborables(inicio_com, fin_com, festivos)
    days_elapsed = contar_dias_laborables(inicio_com, fg, festivos)
    days_elapsed = min(days_elapsed, days_total)

    # KPI global
    pace       = total / days_elapsed if days_elapsed > 0 else 0
    projection = pace * days_total
    budget_cur = BUDGET_26.get(com_month, {}).get("TOTAL", 0)
    var_proy   = round(projection - budget_cur, 2)

    # KPI por centro
    def center_kpi(c):
        r = real.get(c, 0)
        b = BUDGET_26.get(com_month, {}).get(c, 0)
        p = (r / days_elapsed * days_total) if days_elapsed > 0 else 0
        return {
            "real": round(r, 2),
            "budget": b,
            "projection": round(p, 2),
            "pace": round(r / days_elapsed, 2) if days_elapsed > 0 else 0,
            "varProyVsBudget": round(p - b, 2),
            "cumplPct": round(p / b * 100, 1) if b else 0,
            "daysElapsed": days_elapsed,
            "daysTotal": days_total,
        }

    # Histórico total (para rolling)
    hist_total_raw = {}
    for year, months in MONTHLY_BY_CENTER.items():
        y = int(year)
        hist_total_raw[y] = [None] * 12
        for m, rc in months.items():
            hist_total_raw[y][m - 1] = round(sum(rc.values()))
    # Añadir años más antiguos de totales estáticos
    for yr, monthly in TOTAL_MONTHLY.items():
        if yr not in hist_total_raw:
            hist_total_raw[yr] = monthly
    # Mes actual en curso
    hist_total_raw.setdefault(com_year, [None] * 12)
    hist_total_raw[com_year][com_month - 1] = round(total)
    # Convertir a str keys para build functions
    hist_total_str = {str(k): v for k, v in hist_total_raw.items()}

    rolling = build_rolling(hist_total_raw)

    # Histórico por centro
    def extend_hist_center(static: dict, cur_center: str) -> dict:
        h = {k: list(v) for k, v in static.items()}
        # Actualizar meses 2026 completados
        for m_str, rcs in MONTHLY_BY_CENTER.get(str(com_year), {}).items():
            mi = int(m_str) - 1
            yr = str(com_year)
            h.setdefault(yr, [None] * 12)
            h[yr][mi] = round(rcs.get(cur_center, 0))
        # Mes actual parcial
        yr = str(com_year)
        h.setdefault(yr, [None] * 12)
        h[yr][com_month - 1] = round(real.get(cur_center, 0))
        return {k: v for k, v in sorted(h.items())}

    hist_central  = extend_hist_center(HIST_CENTRAL, "CENTRAL")
    hist_alcarras = extend_hist_center(HIST_ALCARRAS, "ALCARRAS")
    hist_almozara = extend_hist_center(HIST_ALMOZARA, "ALMOZARA")
    hist_corona   = extend_hist_center(HIST_CORONA, "CORONA")

    # Series mensuales
    monthly_series = build_monthly_series(
        com_year, com_month, {**real, "TOTAL": total}, projection
    )

    # ── Ensamblar JSON ─────────────────────────────────────────────────────────
    data = {
        "meta": {
            "lastLoadDate":       fg.isoformat(),
            "lastRunTs":          datetime.utcnow().isoformat(timespec="seconds"),
            "comercialMonth":     f"{com_year}-{com_month:02d}",
            "comercialMonthLabel": f"{'Enero Febrero Marzo Abril Mayo Junio Julio Agosto Septiembre Octubre Noviembre Diciembre'.split()[com_month-1]} {com_year}",
        },
        "kpis": {
            "global": {
                "real":          round(total, 2),
                "budget":        budget_cur,
                "projection":    round(projection, 2),
                "pace":          round(pace, 2),
                "varProyVsBudget": var_proy,
                "pctProyBudget": round(projection / budget_cur * 100, 1) if budget_cur else 0,
                "daysElapsed":   days_elapsed,
                "daysTotal":     days_total,
            },
            "byCenter": {c: center_kpi(c) for c in ["CENTRAL","ALCARRAS","ALMOZARA","CORONA"]},
        },
        "series": {
            "monthly": monthly_series,
            "rolling": rolling,
        },
        "vendors": {
            c: vend.get(c, []) for c in ["CENTRAL","ALCARRAS","ALMOZARA","CORONA"]
        },
        "historical": {
            "TOTAL":    build_hist_total(com_year, com_month,
                            {com_year: {com_month: {c: real.get(c,0) for c in real if c!="TOTAL"}}}),
            "CENTRAL":  hist_central,
            "ALCARRAS": hist_alcarras,
            "ALMOZARA": hist_almozara,
            "CORONA":   hist_corona,
        },
    }

    Path(output_path).write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"✓ {output_path} escrito  ({len(json.dumps(data))//1024} KB)")
    return data


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=== Extractor Europa – inicio ===")

    # 1. Descargar Excel ERP desde Gmail
    try:
        payload, fname, uid, fecha_gen = fetch_latest_erp_excel()
        print(f"✓ Excel descargado: {fname}  (fecha generación: {fecha_gen})")
    except Exception as e:
        print(f"✗ Error IMAP: {e}")
        traceback.print_exc()
        raise SystemExit(1)

    # 2. Guardar temporalmente y parsear
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tf.write(payload)
        tmp_path = tf.name

    try:
        erp_data = parsear_excel_erp(tmp_path)
        print(f"✓ ERP parseado: total={erp_data['total']:.2f} €")
        print(f"  Canales: {erp_data['por_canal']}")
        print(f"  Vendedores: { {c: len(v) for c,v in erp_data['por_vendedor'].items()} }")
    finally:
        os.unlink(tmp_path)

    # 3. Guardar en Google Sheets (idempotente)
    try:
        append_to_sheets(erp_data)
    except Exception as e:
        print(f"  Sheets warning: {e}")

    # 4. Generar data.json
    generar_data_json(erp_data, output_path="data.json")

    print("=== Extractor Europa – fin ===")


if __name__ == "__main__":
    main()
